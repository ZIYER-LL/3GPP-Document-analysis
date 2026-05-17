from openpyxl import load_workbook


def get_cell_value(cell):
    return None if cell is None or cell.value is None else str(cell.value).strip()


def looks_like_url(text):
    return isinstance(text, str) and text.startswith(("http://", "https://"))


def get_hyperlink_target(cell):
    hyperlink = cell.hyperlink

    if isinstance(hyperlink, str) and looks_like_url(hyperlink):
        return hyperlink.strip()

    if hyperlink is not None:
        target = getattr(hyperlink, "target", None)
        if isinstance(target, str) and looks_like_url(target):
            return target.strip()

    if looks_like_url(cell.value):
        return str(cell.value).strip()

    return None


def normalize_header(value):
    if value is None:
        return None
    return str(value).strip().lower()


def build_headers(ws):
    headers = {}
    for cell in ws[1]:
        value = normalize_header(cell.value)
        if value:
            headers[value] = cell.column
    return headers


def safe_cell(ws, row_idx, headers, header_name):
    col = headers.get(normalize_header(header_name))
    if not col:
        return None
    return ws.cell(row=row_idx, column=col)


def safe_value(ws, row_idx, headers, header_name):
    return get_cell_value(safe_cell(ws, row_idx, headers, header_name))


def import_tdoc_sheet(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = build_headers(ws)
    print("Detected headers:", headers)

    required = ["tdoc", "title"]
    missing = [x for x in required if x not in headers]
    if missing:
        raise ValueError(f"缺少必要表头: {missing}，当前表头: {list(headers.keys())}")

    records = []

    for row_idx in range(2, ws.max_row + 1):
        tdoc_cell = safe_cell(ws, row_idx, headers, "TDoc")
        revised_to_cell = safe_cell(ws, row_idx, headers, "Revised to")

        record = {
            "tdoc_id": get_cell_value(tdoc_cell),
            "tdoc_url": get_hyperlink_target(tdoc_cell),

            "title": safe_value(ws, row_idx, headers, "Title"),
            "source": safe_value(ws, row_idx, headers, "Source"),
            "contact": safe_value(ws, row_idx, headers, "Contact"),
            "type_raw": safe_value(ws, row_idx, headers, "Type"),
            "for_raw": safe_value(ws, row_idx, headers, "For"),
            "abstract": safe_value(ws, row_idx, headers, "Abstract"),

            "agenda": safe_value(ws, row_idx, headers, "Agenda"),
            "agenda_item": safe_value(ws, row_idx, headers, "Agenda Item"),
            "agenda_item_desc": safe_value(ws, row_idx, headers, "Agenda item description"),

            "tdoc_status": safe_value(ws, row_idx, headers, "TDoc Status"),
            "uploaded_raw": safe_value(ws, row_idx, headers, "Uploaded"),

            "is_revision_of": safe_value(ws, row_idx, headers, "Is revision of"),
            "revised_to": get_cell_value(revised_to_cell),
            "revised_to_url": get_hyperlink_target(revised_to_cell),

            "release": safe_value(ws, row_idx, headers, "Release"),
            "spec": safe_value(ws, row_idx, headers, "Spec"),
            "version": safe_value(ws, row_idx, headers, "Version"),
            "related_wi": safe_value(ws, row_idx, headers, "Related WI"),

            "cr": safe_value(ws, row_idx, headers, "CR"),
            "cr_rev": safe_value(ws, row_idx, headers, "CR rev"),

            "reply_to": safe_value(ws, row_idx, headers, "Reply to"),
            "to_field": safe_value(ws, row_idx, headers, "To"),
            "cc_field": safe_value(ws, row_idx, headers, "CC"),
            "original": safe_value(ws, row_idx, headers, "Original"),
        }

        if record["tdoc_id"] or record["title"]:
            records.append(record)

    return records